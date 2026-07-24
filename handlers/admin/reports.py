"""Reports: on-demand stats, daily/weekly summaries, Excel export with multiple sheets."""
from __future__ import annotations

import io
from datetime import datetime, timezone
from datetime import timedelta

from aiogram import F, Router
from aiogram.types import BufferedInputFile, CallbackQuery
from sqlalchemy import func, select

from database.db import async_session
from database.models import (
    AdminUser,
    BroadcastLog,
    Category,
    Material,
    Order,
    OrderItem,
    Product,
    PromoCode,
    Review,
    Size,
    User,
    UserTagHistory,
)
from keyboards.admin_kb import admin_back_kb, admin_stats_kb
from services.reports import (
    build_daily_report,
    build_weekly_report,
    dropoff_stats,
    funnel_stats,
    general_stats,
    top_products,
)
from utils.security import IsAdminFilter

router = Router()
router.callback_query.filter(IsAdminFilter())


async def _get_perms(telegram_id: int) -> set[str]:
    from utils.security import get_admin_role_async, _role_permissions
    role = await get_admin_role_async(telegram_id)
    return _role_permissions().get(role, set()) if role else set()


@router.callback_query(F.data == "admin:stats")
async def cb_admin_stats(callback: CallbackQuery):
    perms = await _get_perms(callback.from_user.id)
    if "stats" not in perms:
        await callback.answer("Ruxsat yo'q.", show_alert=True)
        return
    await callback.message.edit_text(
        "📊 <b>Statistika</b>\n\nBo'limni tanlang:",
        reply_markup=admin_stats_kb(),
    )
    await callback.answer()


@router.callback_query(F.data == "stats:general")
async def cb_stats_general(callback: CallbackQuery):
    perms = await _get_perms(callback.from_user.id)
    if "stats" not in perms:
        await callback.answer("Ruxsat yo'q.", show_alert=True)
        return
    async with async_session() as session:
        text = await general_stats(session)
    await callback.message.edit_text(text, reply_markup=admin_back_kb("admin:stats"))
    await callback.answer()


@router.callback_query(F.data == "stats:dropoff")
async def cb_stats_dropoff(callback: CallbackQuery):
    perms = await _get_perms(callback.from_user.id)
    if "stats" not in perms:
        await callback.answer("Ruxsat yo'q.", show_alert=True)
        return
    async with async_session() as session:
        text = await dropoff_stats(session)
    await callback.message.edit_text(text, reply_markup=admin_back_kb("admin:stats"))
    await callback.answer()


@router.callback_query(F.data == "stats:funnel")
async def cb_stats_funnel(callback: CallbackQuery):
    perms = await _get_perms(callback.from_user.id)
    if "stats" not in perms:
        await callback.answer("Ruxsat yo'q.", show_alert=True)
        return
    async with async_session() as session:
        text = await funnel_stats(session)
    await callback.message.edit_text(text, reply_markup=admin_back_kb("admin:stats"))
    await callback.answer()


@router.callback_query(F.data == "stats:top_products")
async def cb_stats_top(callback: CallbackQuery):
    perms = await _get_perms(callback.from_user.id)
    if "stats" not in perms:
        await callback.answer("Ruxsat yo'q.", show_alert=True)
        return
    async with async_session() as session:
        text = await top_products(session)
    await callback.message.edit_text(text, reply_markup=admin_back_kb("admin:stats"))
    await callback.answer()


@router.callback_query(F.data == "stats:daily_now")
async def cb_stats_daily_now(callback: CallbackQuery):
    perms = await _get_perms(callback.from_user.id)
    if "stats" not in perms:
        await callback.answer("Ruxsat yo'q.", show_alert=True)
        return
    async with async_session() as session:
        text = await build_daily_report(session, target_date=datetime.now(timezone.utc))
    await callback.message.edit_text(text, reply_markup=admin_back_kb("admin:stats"))
    await callback.answer()


@router.callback_query(F.data == "stats:weekly_now")
async def cb_stats_weekly_now(callback: CallbackQuery):
    perms = await _get_perms(callback.from_user.id)
    if "stats" not in perms:
        await callback.answer("Ruxsat yo'q.", show_alert=True)
        return
    async with async_session() as session:
        text = await build_weekly_report(session)
    await callback.message.edit_text(text, reply_markup=admin_back_kb("admin:stats"))
    await callback.answer()


@router.callback_query(F.data == "admin:export")
async def export_orders_xlsx(callback: CallbackQuery):
    """Generate .xlsx with multiple sheets: Orders, Items, Users, Dropoff, Broadcasts, Reviews."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    perms = await _get_perms(callback.from_user.id)
    if "export" not in perms:
        await callback.answer("Ruxsat yo'q.", show_alert=True)
        return

    wb = Workbook()

    # ---------- Sheet 1: Orders ----------
    ws = wb.active
    ws.title = "Orders"
    headers = [
        "Order #", "Status", "User ID", "Telegram ID", "Username",
        "Full name", "Phone", "Address", "Total price", "Discount",
        "Promo", "Deadline", "Created at", "Finalized at", "Items count",
    ]
    ws.append(headers)
    for col in ws[1]:
        col.font = Font(bold=True)

    async with async_session() as session:
        orders = (
            await session.execute(
                select(Order).where(Order.status != "draft").order_by(Order.created_at.desc())
            )
        ).scalars().all()
        for o in orders:
            user = await session.get(User, o.user_id) if o.user_id else None
            items_count = (
                await session.execute(
                    select(func.count()).select_from(OrderItem).where(OrderItem.order_id == o.id)
                )
            ).scalar_one()
            ws.append([
                o.order_number, o.status, o.user_id,
                user.telegram_id if user else None,
                user.username if user else None,
                user.full_name if user else None,
                user.phone if user else None,
                user.address if user else None,
                o.total_price, o.discount, o.promo_code, o.deadline_type,
                o.created_at.strftime("%Y-%m-%d %H:%M:%S") if o.created_at else None,
                o.finalized_at.strftime("%Y-%m-%d %H:%M:%S") if o.finalized_at else None,
                items_count,
            ])

        # ---------- Sheet 2: Items ----------
        ws2 = wb.create_sheet("Items")
        ws2.append(["Order #", "Item ID", "Product ID", "Product caption",
                    "Material", "Size", "Price"])
        for col in ws2[1]:
            col.font = Font(bold=True)
        for o in orders:
            res = await session.execute(
                select(OrderItem, Product, Material, Size)
                .join(Product, OrderItem.product_id == Product.id, isouter=True)
                .join(Material, OrderItem.material_id == Material.id, isouter=True)
                .join(Size, OrderItem.size_id == Size.id, isouter=True)
                .where(OrderItem.order_id == o.id)
            )
            for item, prod, mat, sz in res.all():
                ws2.append([
                    o.order_number, item.id, item.product_id,
                    prod.caption_uz if prod else None,
                    mat.name_uz if mat else None,
                    sz.name_uz if sz else None,
                    item.price,
                ])

        # ---------- Sheet 3: Users ----------
        ws3 = wb.create_sheet("Users")
        ws3.append(["ID", "Telegram ID", "Username", "Full name", "Phone", "Address",
                    "Tag", "Language", "First seen", "Last active"])
        for col in ws3[1]:
            col.font = Font(bold=True)
        users = (await session.execute(select(User).order_by(User.id))).scalars().all()
        for u in users:
            ws3.append([
                u.id, u.telegram_id, u.username, u.full_name, u.phone, u.address,
                u.tag, u.language,
                u.first_seen_at.strftime("%Y-%m-%d %H:%M:%S") if u.first_seen_at else None,
                u.last_active_at.strftime("%Y-%m-%d %H:%M:%S") if u.last_active_at else None,
            ])

        # ---------- Sheet 4: Drop-off (current tags) ----------
        ws4 = wb.create_sheet("Dropoff")
        ws4.append(["Tag", "Count"])
        for col in ws4[1]:
            col.font = Font(bold=True)
        rows = (await session.execute(
            select(User.tag, func.count()).group_by(User.tag).order_by(func.count().desc())
        )).all()
        for tag, cnt in rows:
            ws4.append([tag, cnt])

        # ---------- Sheet 5: Broadcasts ----------
        ws5 = wb.create_sheet("Broadcasts")
        ws5.append(["ID", "Admin ID", "Target", "Message", "Sent", "Failed",
                    "Blocked IDs", "Sent at"])
        for col in ws5[1]:
            col.font = Font(bold=True)
        bcs = (await session.execute(select(BroadcastLog).order_by(BroadcastLog.id.desc()))).scalars().all()
        for b in bcs:
            ws5.append([
                b.id, b.admin_id, b.target_tag, b.message, b.sent_count,
                b.failed_count, b.blocked_user_ids,
                b.sent_at.strftime("%Y-%m-%d %H:%M:%S") if b.sent_at else None,
            ])

        # ---------- Sheet 6: Reviews ----------
        ws6 = wb.create_sheet("Reviews")
        ws6.append(["Review ID", "Order #", "User Telegram ID", "Rating", "Comment", "Created at"])
        for col in ws6[1]:
            col.font = Font(bold=True)
        reviews_q = await session.execute(
            select(Review, Order, User)
            .join(Order, Review.order_id == Order.id)
            .join(User, Review.user_id == User.id)
            .order_by(Review.created_at.desc())
        )
        for r, o, u in reviews_q.all():
            ws6.append([
                r.id, o.order_number, u.telegram_id, r.rating, r.comment,
                r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else None,
            ])

        # ---------- Sheet 7: Promos ----------
        ws7 = wb.create_sheet("Promos")
        ws7.append(["Code", "Type", "Value", "Min order", "Max uses", "Uses",
                    "Valid from", "Valid until", "Active", "Created at"])
        for col in ws7[1]:
            col.font = Font(bold=True)
        promos = (await session.execute(select(PromoCode).order_by(PromoCode.id.desc()))).scalars().all()
        for p in promos:
            ws7.append([
                p.code, p.discount_type, p.discount_value, p.min_order_amount,
                p.max_uses, p.uses_count,
                p.valid_from.strftime("%Y-%m-%d %H:%M:%S") if p.valid_from else None,
                p.valid_until.strftime("%Y-%m-%d %H:%M:%S") if p.valid_until else None,
                p.is_active,
                p.created_at.strftime("%Y-%m-%d %H:%M:%S") if p.created_at else None,
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await callback.message.answer_document(
        document=BufferedInputFile(buf.read(), filename="full_export.xlsx"),
        caption="📋 To'liq eksport (.xlsx, 7 varaq)",
        reply_markup=admin_back_kb("admin:panel"),
    )
    await callback.answer("Eksport tayyor 📄")
